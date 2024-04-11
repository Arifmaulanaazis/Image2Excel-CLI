import argparse
import concurrent.futures
import matplotlib.colors
import numpy as np
import openpyxl
import tempfile, os
from PIL import Image
from tqdm import tqdm


def process_pixel(row, col):
    pixel = image_array[row, col]
    rgb_tuple = pixel[:3]
    hexa = matplotlib.colors.rgb2hex([1.0 * clr / 255 for clr in rgb_tuple])
    colour = openpyxl.styles.colors.Color(rgb=f"FF{hexa.lstrip('#')}")
    worksheet.cell(row=row + 2, column=col + 2).fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=colour)


print('Image2Excel || Mengubah gambar menjadi file Excel || Copyright Arif Maulana 2023\n')
parser = argparse.ArgumentParser(description='Image2Excel || Mengubah gambar menjadi file Excel || Copyright Arif Maulana 2023')
parser.add_argument('--input', '-i', '-I', type=str, help='Path input gambar', required=True)
parser.add_argument('--downscale', '-d', '-D',  type=int, help='Rasio downscale ukuran gambar', default=10)
parser.add_argument('--output', '-o', '-O',  type=str, help='Path output file Excel', required=True)
args = parser.parse_args()




input_gambar = args.input
downscale_berapa_kali = args.downscale
nama_output = args.output

if input_gambar == None:
    print('Path input gambar')
    exit()

if nama_output == None:
    print('Path output file Excel')
    exit()

if os.path.exists(input_gambar) == False:
    print(f'File {input_gambar} tidak ditemukan')
    exit()

try:
    downscale_berapa_kali = int(downscale_berapa_kali)
except:
    print('Rasio downscale harus berupa angka')
    exit()

if nama_output.endswith('.xlsx') == False:
    nama_output = nama_output + '.xlsx'


workbook = openpyxl.load_workbook("./data/data.xlsx")
worksheet = workbook.active
image = Image.open(input_gambar)
image = image.reduce(downscale_berapa_kali)
image_array = np.asarray(image)
total_pixels = image_array.shape[0] * image_array.shape[1]
with tqdm(total=total_pixels, desc="memproses", unit="pixel") as pbar:
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = []
        for row in range(image_array.shape[0]):
            for col in range(image_array.shape[1]):
                future = executor.submit(process_pixel, row, col)
                future.add_done_callback(lambda x: pbar.update(1))
                futures.append(future)

        concurrent.futures.wait(futures)

workbook.save(nama_output)

print("SELESAI")
print(f'Berhasil menyimpan file {nama_output}')